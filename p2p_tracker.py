import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import numpy as np
import json
import os
import io
import glob
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import openpyxl.cell.cell

# Set page config
st.set_page_config(
    page_title="P2P & E-Voucher Tracker",
    page_icon="💱",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Add custom CSS
st.markdown("""
<style>
    .main {
        padding: 1rem;
    }
    .stApp {
        max-width: 1200px;
        margin: 0 auto;
    }
    h1, h2, h3 {
        color: #1E88E5;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 24px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 50px;
        white-space: pre-wrap;
        background-color: #F0F2F6;
        border-radius: 4px 4px 0px 0px;
        gap: 1px;
        padding-top: 10px;
        padding-bottom: 10px;
    }
    .stTabs [aria-selected="true"] {
        background-color: #1E88E5;
        color: white;
    }
    .currency-card {
        background-color: #f8f9fa;
        border-radius: 10px;
        padding: 20px;
        margin-bottom: 20px;
        border-left: 5px solid #1E88E5;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }
    .currency-title {
        font-size: 18px;
        font-weight: bold;
        margin-bottom: 10px;
        color: #333;
    }
    .currency-value {
        font-size: 24px;
        font-weight: bold;
        color: #1E88E5;
    }
    .currency-subtitle {
        font-size: 14px;
        color: #666;
        margin-top: 5px;
    }
    .total-value-card {
        background-color: #e3f2fd;
        border-radius: 10px;
        padding: 20px;
        margin-top: 20px;
        border-left: 5px solid #1565C0;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }
    .profit-card {
        background-color: #e8f5e9;
        border-radius: 10px;
        padding: 20px;
        margin-bottom: 20px;
        border-left: 5px solid #2E7D32;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }
    .profit-title {
        font-size: 18px;
        font-weight: bold;
        margin-bottom: 10px;
        color: #2E7D32;
    }
    .profit-value {
        font-size: 24px;
        font-weight: bold;
        color: #2E7D32;
    }
</style>
""", unsafe_allow_html=True)

# Title
st.title("💱 P2P & E-Voucher Transaction Tracker")
st.markdown("---")

# Sidebar for inputs
st.sidebar.header("Initial Settings")

# Function to load and process data
def process_data(file, initial_balances, initial_usdt_rate):
    try:
        # Read Excel file
        df = pd.read_excel(file)

        # Check if required columns exist
        required_columns = ['Date', 'Type', 'TradeType', 'Currency', 'Real Amount', 'USDT', 'Price', 'Status']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            st.error(f"Missing required columns: {', '.join(missing_columns)}")
            return None

        # Add AED/EGP column if it doesn't exist
        if 'AED/EGP' not in df.columns:
            df['AED/EGP'] = 13.75  # Default value

        # Filter completed transactions
        df = df[df['Status'] == 'COMPLETED'].copy()

        # Convert date to datetime
        try:
            df['Date'] = pd.to_datetime(df['Date'], format='%d/%m/%Y, %H:%M:%S')
        except:
            try:
                # Try alternative format
                df['Date'] = pd.to_datetime(df['Date'])
            except:
                st.error("Could not parse date format. Please ensure dates are in format 'DD/MM/YYYY, HH:MM:SS'")
                return None

        # Sort by date
        df = df.sort_values('Date')

        # Initialize balances dictionary with initial balances
        balances = initial_balances.copy()

        # Initialize lists to store balance history
        balance_history = []
        profit_history = []
        cash_flow = []

        # Initialize FIFO queue for USDT purchases with initial USDT
        usdt_queue = []

        # Add initial USDT to the queue with the initial rate
        if 'USDT' in initial_balances and initial_balances['USDT'] > 0:
            usdt_queue.append((initial_balances['USDT'], initial_usdt_rate))

        # Initialize profit tracking
        total_p2p_profit = 0
        total_evoucher_profit = 0

        # Calculate average USDT rate
        def calculate_avg_usdt_rate():
            if not usdt_queue or sum(amount for amount, _ in usdt_queue) == 0:
                return initial_usdt_rate

            total_usdt = sum(amount for amount, _ in usdt_queue)
            total_value = sum(amount * rate for amount, rate in usdt_queue)
            return total_value / total_usdt if total_usdt > 0 else initial_usdt_rate

        # Process each transaction
        for idx, row in df.iterrows():
            date = row['Date']
            trade_type = row['TradeType']
            transaction_type = row['Type']
            currency = row['Currency']
            real_amount = row['Real Amount']
            usdt_amount = row['USDT']
            price = row['Price']
            aed_egp_rate = row['AED/EGP']

            # Ensure currency exists in balances
            if currency not in balances:
                balances[currency] = 0

            # Calculate transaction profit
            transaction_profit = 0

            # Process based on trade type
            if trade_type == 'P2P':
                # P2P transactions
                if transaction_type == 'Buy':
                    # Buy USDT: decrease local currency, increase USDT
                    balances[currency] -= real_amount
                    balances['USDT'] += usdt_amount

                    # Add to USDT queue with purchase price in AED
                    if currency == 'AED':
                        usdt_queue.append((usdt_amount, price))
                    else:
                        # Convert price to AED equivalent
                        aed_price = price / aed_egp_rate if currency == 'EGP' else price
                        usdt_queue.append((usdt_amount, aed_price))

                elif transaction_type == 'Sell':
                    # Sell USDT: increase local currency, decrease USDT
                    balances[currency] += real_amount
                    balances['USDT'] -= usdt_amount

                    # Calculate profit using FIFO method
                    remaining_usdt = usdt_amount
                    sell_price_in_aed = price

                    # For debugging
                    debug_info = []

                    # Convert sell price to AED if needed
                    if currency != 'AED':
                        sell_price_in_aed = price * aed_egp_rate if currency == 'EGP' else price

                    # Calculate total sell value
                    total_sell_value = sell_price_in_aed * usdt_amount
                    total_cost_value = 0

                    # Track original purchase rate for this sell transaction
                    original_purchase_rate = 0

                    # Track USDT used for detailed calculation
                    usdt_used_details = []

                    while remaining_usdt > 0 and usdt_queue:
                        bought_usdt, bought_price = usdt_queue[0]

                        if bought_usdt <= remaining_usdt:
                            # Use entire bought batch
                            batch_cost = bought_usdt * bought_price
                            total_cost_value += batch_cost

                            # Add to debug info
                            debug_info.append(f"Used {bought_usdt} USDT @ {bought_price} = {batch_cost} AED")
                            usdt_used_details.append((bought_usdt, bought_price))

                            remaining_usdt -= bought_usdt
                            usdt_queue.pop(0)
                        else:
                            # Use partial bought batch
                            batch_cost = remaining_usdt * bought_price
                            total_cost_value += batch_cost

                            # Add to debug info
                            debug_info.append(f"Used {remaining_usdt} USDT @ {bought_price} = {batch_cost} AED")
                            usdt_used_details.append((remaining_usdt, bought_price))

                            usdt_queue[0] = (bought_usdt - remaining_usdt, bought_price)
                            remaining_usdt = 0

                    # Calculate profit as the difference between sell value and cost
                    transaction_profit = total_sell_value - total_cost_value

                    # Calculate original purchase rate (weighted average of used USDT)
                    if usdt_amount > 0:
                        original_purchase_rate = total_cost_value / usdt_amount

                    # Add detailed calculation to debug info
                    debug_info.append(f"Sell value: {usdt_amount} USDT @ {sell_price_in_aed} = {total_sell_value} AED")
                    debug_info.append(f"Cost value: {total_cost_value} AED")
                    debug_info.append(f"Original purchase rate: {original_purchase_rate} AED")
                    debug_info.append(f"Profit: {transaction_profit} AED")

                    # Store debug info for this transaction
                    if 'debug_transactions' not in globals():
                        globals()['debug_transactions'] = []

                    globals()['debug_transactions'].append({
                        'date': date,
                        'type': 'Sell',
                        'usdt_amount': usdt_amount,
                        'sell_price': sell_price_in_aed,
                        'total_sell_value': total_sell_value,
                        'usdt_used': usdt_used_details,
                        'total_cost_value': total_cost_value,
                        'original_purchase_rate': original_purchase_rate,
                        'profit': transaction_profit,
                        'debug_info': debug_info
                    })

                    total_p2p_profit += transaction_profit

            elif trade_type == 'E-Voucher':
                # E-Voucher transactions
                if transaction_type == 'Buy':
                    # Receiving AED from workers
                    balances['AED'] += real_amount

                elif transaction_type == 'Sell':
                    # Selling USDT for EGP to families
                    # Note: We don't add EGP to balances as it's sent to families, not kept
                    balances['USDT'] -= usdt_amount

                    # Calculate cost of USDT in AED
                    usdt_cost_in_aed = 0
                    remaining_usdt = usdt_amount

                    while remaining_usdt > 0 and usdt_queue:
                        bought_usdt, bought_price = usdt_queue[0]

                        if bought_usdt <= remaining_usdt:
                            # Use entire bought batch
                            usdt_cost_in_aed += bought_usdt * bought_price
                            remaining_usdt -= bought_usdt
                            usdt_queue.pop(0)
                        else:
                            # Use partial bought batch
                            usdt_cost_in_aed += remaining_usdt * bought_price
                            usdt_queue[0] = (bought_usdt - remaining_usdt, bought_price)
                            remaining_usdt = 0

                    # For E-Voucher, profit is calculated separately
                    # We don't add it to the P2P profit calculation
                    # The profit is the AED received (from Buy transactions) minus the cost of USDT sold
                    # This will be calculated separately in the E-Voucher analysis
                    transaction_profit = 0  # Don't calculate profit here

            # Calculate current average USDT rate
            current_avg_usdt_rate = calculate_avg_usdt_rate()

            # Calculate USDT value in AED
            usdt_value_in_aed = balances.get('USDT', 0) * current_avg_usdt_rate

            # Record balance history
            balance_record = {
                'Date': date,
                'TradeType': trade_type,
                'Type': transaction_type,
                'Currency': currency,
                'Real Amount': real_amount,
                'USDT': usdt_amount,
                'Price': price,
                'Profit': transaction_profit if trade_type == 'P2P' else 0,  # Only show profit for P2P transactions
                'Current USDT Rate': current_avg_usdt_rate,
                'USDT Value (AED)': usdt_value_in_aed
            }

            # Add current balances to record
            for curr, bal in balances.items():
                balance_record[f'{curr} Balance'] = bal

            balance_history.append(balance_record)

            # Record profit history
            profit_history.append({
                'Date': date,
                'P2P Profit': total_p2p_profit,
                'E-Voucher Profit': total_evoucher_profit,
                'Total Profit': total_p2p_profit + total_evoucher_profit
            })

            # Record cash flow
            cash_flow_record = {
                'Date': date,
                'Transaction': f"{transaction_type} {usdt_amount} USDT @ {price} ({trade_type})",
                'AED Balance': balances.get('AED', 0),
                'USDT Balance': balances.get('USDT', 0),
                'USDT Rate': current_avg_usdt_rate,
                'USDT Value (AED)': usdt_value_in_aed,
                'Total Value (AED)': balances.get('AED', 0) + usdt_value_in_aed,  # Only count AED and USDT value
                'Transaction Profit': transaction_profit if trade_type == 'P2P' else 0,  # Only show profit for P2P transactions
                'Cumulative P2P Profit': total_p2p_profit
            }

            # Add original purchase rate for sell transactions
            if transaction_type == 'Sell' and trade_type == 'P2P':
                cash_flow_record['Original Purchase Rate'] = original_purchase_rate
            cash_flow.append(cash_flow_record)

        # Create DataFrame from balance history
        balance_df = pd.DataFrame(balance_history)
        profit_df = pd.DataFrame(profit_history)
        cash_flow_df = pd.DataFrame(cash_flow)

        # Calculate final USDT rate
        final_usdt_rate = calculate_avg_usdt_rate()

        # Calculate final USDT value in AED
        final_usdt_value = balances.get('USDT', 0) * final_usdt_rate

        # Create debug transactions DataFrame if available
        debug_df = None
        if 'debug_transactions' in globals() and globals()['debug_transactions']:
            debug_df = pd.DataFrame(globals()['debug_transactions'])

        # Calculate E-Voucher profit separately
        evoucher_profit = 0
        evoucher_aed_received = 0  # Initialize evoucher_aed_received
        evoucher_df = df[df['TradeType'] == 'E-Voucher'].copy()

        if not evoucher_df.empty:
            # Get all Buy transactions (AED received)
            buy_transactions = evoucher_df[evoucher_df['Type'] == 'Buy']
            total_aed_received = buy_transactions['Real Amount'].sum()
            evoucher_aed_received = total_aed_received  # Store the total AED received

            # Get all Sell transactions (USDT sold)
            sell_transactions = evoucher_df[evoucher_df['Type'] == 'Sell']

            # Calculate the cost of USDT sold based on the current USDT rate
            total_usdt_cost = 0
            for _, row in sell_transactions.iterrows():
                usdt_amount = row['USDT']
                # Use the average USDT rate at the time of the transaction
                idx = balance_df[balance_df['Date'] == row['Date']].index[0]
                usdt_rate = balance_df.loc[idx, 'Current USDT Rate']
                total_usdt_cost += usdt_amount * usdt_rate

            # Calculate profit
            evoucher_profit = total_aed_received - total_usdt_cost

        return {
            'transactions': df,
            'balance_history': balance_df,
            'profit_history': profit_df,
            'cash_flow': cash_flow_df,
            'final_balances': balances,
            'initial_balances': initial_balances,  # Add initial balances
            'initial_usdt_rate': initial_usdt_rate,  # Add initial USDT rate
            'total_p2p_profit': total_p2p_profit,
            'total_evoucher_profit': evoucher_profit,  # Use the separately calculated E-Voucher profit
            'evoucher_aed_received': evoucher_aed_received,  # Add E-Voucher AED received
            'final_usdt_rate': final_usdt_rate,
            'final_usdt_value': final_usdt_value,
            'debug_transactions': debug_df  # Add debug transactions
        }

    except Exception as e:
        st.error(f"Error processing file: {str(e)}")
        return None

# Function to create balance chart
def create_balance_chart(balance_df, currencies):
    fig = go.Figure()

    for currency in currencies:
        fig.add_trace(
            go.Scatter(
                x=balance_df['Date'],
                y=balance_df[f'{currency} Balance'],
                mode='lines+markers',
                name=currency
            )
        )

    fig.update_layout(
        title='Balance History',
        xaxis_title='Date',
        yaxis_title='Balance',
        legend_title='Currency',
        height=500
    )

    return fig

# Function to create profit chart
def create_profit_chart(profit_df):
    fig = go.Figure()

    fig.add_trace(
        go.Scatter(
            x=profit_df['Date'],
            y=profit_df['P2P Profit'],
            mode='lines+markers',
            name='P2P Profit'
        )
    )

    fig.add_trace(
        go.Scatter(
            x=profit_df['Date'],
            y=profit_df['E-Voucher Profit'],
            mode='lines+markers',
            name='E-Voucher Profit'
        )
    )

    fig.add_trace(
        go.Scatter(
            x=profit_df['Date'],
            y=profit_df['Total Profit'],
            mode='lines+markers',
            name='Total Profit',
            line=dict(width=3)
        )
    )

    fig.update_layout(
        title='Profit History',
        xaxis_title='Date',
        yaxis_title='Profit (AED)',
        legend_title='Profit Type',
        height=500
    )

    return fig

# Function to export all data to Excel
def export_to_excel(results, manual_aed_received=0):
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # Define border style for tables
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

        # Create a BytesIO object to store the Excel file
        output = io.BytesIO()

        # Create a workbook
        wb = Workbook()

        # Use the default sheet
        ws = wb.active
        ws.title = "P2P Tracker Data"

        # Define border style for tables
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

        # Get transaction date for the title
        transaction_date = "No Data"
        if not results['cash_flow'].empty:
            try:
                # Get only the first date of transactions
                first_date = results['cash_flow']['Date'].iloc[0] if 'Date' in results['cash_flow'].columns else ""

                # Make sure date is string
                first_date = str(first_date)

                # Use just the date part for the title
                transaction_date = first_date

            except Exception as e:
                # If there's any error, use a default
                from datetime import datetime
                today = datetime.now().strftime("%Y-%m-%d")
                transaction_date = today
                print(f"Could not format date for title: {str(e)}")

        # Get total profit for title
        total_profit = results['total_p2p_profit'] + results['total_evoucher_profit']

        # Add professional title with date and profit
        ws['A1'] = f'P2P & E-Voucher Financial Report - {transaction_date}'
        ws['A1'].font = Font(size=12, bold=True)
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center')

        # Add subtitle with profit information
        row = 2
        ws[f'A{row}'] = f'Total Profit: {total_profit:.2f} AED'
        ws[f'A{row}'].font = Font(size=10, bold=True)
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        # Increment row for next section
        row += 1

        # Add Initial Balances section first
        ws[f'A{row}'] = 'INITIAL BALANCES'
        ws[f'A{row}'].font = Font(size=10, bold=True)
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws[f'A{row}'].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        # Add subtitle explaining the table
        row += 1
        ws[f'A{row}'] = 'Opening balances and rates at the beginning of the period'
        ws[f'A{row}'].font = Font(size=8, italic=True)
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        row += 1  # Reduce space

        # Add headers for initial balances
        ws.cell(row=row, column=1, value="Currency").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Initial Balance").font = Font(bold=True)
        ws.cell(row=row, column=3, value="Value (AED)").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        ws.cell(row=row, column=2).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        ws.cell(row=row, column=3).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        row += 1
        initial_balances_start = row

        # Add initial balances data
        aed_balance = 0
        usdt_balance = 0
        usdt_in_aed = 0

        for currency, balance in results.get('initial_balances', {}).items():
            ws.cell(row=row, column=1, value=currency)
            ws.cell(row=row, column=2, value=f"{float(balance):.5f}")

            # Calculate AED value
            if currency == 'AED':
                aed_value = balance
                aed_balance = aed_value
                ws.cell(row=row, column=3, value=f"{aed_value:.5f}")
            elif currency == 'USDT':
                usdt_balance = balance
                usdt_rate = results.get('initial_usdt_rate', 0)
                usdt_in_aed = balance * usdt_rate
                ws.cell(row=row, column=3, value=f"{usdt_in_aed:.5f}")
            else:
                ws.cell(row=row, column=3, value="-")

            row += 1

        # Add initial USDT rate
        ws.cell(row=row, column=1, value="Initial USDT Rate (AED)")
        ws.cell(row=row, column=2, value=f"{results.get('initial_usdt_rate', 0):.5f}")
        ws.cell(row=row, column=3, value="-")
        row += 1

        # Add total AED value row
        total_initial_aed = aed_balance + usdt_in_aed
        ws.cell(row=row, column=1, value="Total Value (AED)")
        ws.cell(row=row, column=2, value="-")
        ws.cell(row=row, column=3, value=f"{total_initial_aed:.5f}")
        ws.cell(row=row, column=3).font = Font(bold=True)

        initial_balances_end = row

        # Add border to the initial balances table
        for r in range(initial_balances_start - 1, initial_balances_end + 1):
            for c in range(1, 4):  # Extended to include the AED value column
                ws.cell(row=r, column=c).border = thin_border

        # Then add Cash Flow table
        row += 2  # Add space
        ws[f'A{row}'] = 'CASH FLOW'
        ws[f'A{row}'].font = Font(size=10, bold=True)
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws[f'A{row}'].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        # Add subtitle explaining the table
        row += 1
        ws[f'A{row}'] = 'Summary of all transactions showing balances and values after each transaction'
        ws[f'A{row}'].font = Font(size=8, italic=True)
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        row += 1  # Reduce space

        # Add Cash Flow data
        if not results['cash_flow'].empty:
            # Create a simplified version of cash flow for better printing
            cash_flow_df = results['cash_flow'].copy()

            # Select only the most important columns for printing
            important_columns = ['Date', 'Transaction', 'AED Balance', 'USDT Balance', 'USDT Rate', 'Total Value (AED)', 'Transaction Profit']

            # Filter columns that exist in the dataframe
            print_columns = [col for col in important_columns if col in cash_flow_df.columns]

            # Create a simplified dataframe
            simplified_cf = cash_flow_df[print_columns].copy()

            # Format numeric columns to 2 decimal places for better readability
            numeric_columns = simplified_cf.select_dtypes(include=['float64', 'int64']).columns

            # Add headers
            for c_idx, col_name in enumerate(simplified_cf.columns, 1):
                cell = ws.cell(row=row, column=c_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            row += 1
            cash_flow_data_start = row

            # Add data rows
            for _, data_row in simplified_cf.iterrows():
                for c_idx, (col_name, value) in enumerate(zip(simplified_cf.columns, data_row), 1):
                    # Handle complex data types
                    if isinstance(value, (list, tuple, dict)):
                        value = str(value)

                    # Format numeric values with 5 decimal places for higher precision
                    if col_name in numeric_columns:
                        try:
                            formatted_value = f"{float(value):.5f}" if value is not None else value
                            cell = ws.cell(row=row, column=c_idx, value=formatted_value)
                        except (ValueError, TypeError):
                            cell = ws.cell(row=row, column=c_idx, value=value)
                    else:
                        cell = ws.cell(row=row, column=c_idx, value=value)

                    # Set alignment and wrap text for Transaction column
                    if col_name == 'Date':
                        cell.alignment = Alignment(horizontal='left')
                    elif col_name == 'Transaction':
                        # Format transaction text to be more readable
                        if isinstance(value, str) and len(value) > 15:
                            # Insert line breaks for common transaction types
                            value = value.replace(" - ", "\n")
                            value = value.replace(" @ ", "\n@ ")

                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        cell.value = value
                        # Set row height to accommodate wrapped text
                        if isinstance(value, str) and "\n" in value:
                            lines = value.count("\n") + 1
                            ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 15, 15 * lines)
                    else:
                        cell.alignment = Alignment(horizontal='right')

                row += 1

            cash_flow_end_row = row - 1

            # Add border to the table
            for r in range(cash_flow_data_start - 1, cash_flow_end_row + 1):
                for c in range(1, len(simplified_cf.columns) + 1):
                    ws.cell(row=r, column=c).border = thin_border

        # Add Final Balances section
        row += 2  # Reduce space
        ws[f'A{row}'] = 'FINAL BALANCES'
        ws[f'A{row}'].font = Font(size=10, bold=True)
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws[f'A{row}'].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        # Add subtitle explaining the table
        row += 1
        ws[f'A{row}'] = 'Closing balances and rates at the end of the period'
        ws[f'A{row}'].font = Font(size=8, italic=True)
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        row += 1  # Reduce space

        # Add headers for final balances
        ws.cell(row=row, column=1, value="Currency").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Final Balance").font = Font(bold=True)
        ws.cell(row=row, column=3, value="Value (AED)").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        ws.cell(row=row, column=2).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        ws.cell(row=row, column=3).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        row += 1
        final_balances_start = row

        # Add final balances data
        aed_balance = 0
        usdt_balance = 0
        usdt_in_aed = 0

        for currency, balance in results['final_balances'].items():
            ws.cell(row=row, column=1, value=currency)
            ws.cell(row=row, column=2, value=f"{balance:.5f}")

            # Calculate AED value
            if currency == 'AED':
                aed_value = balance
                aed_balance = aed_value
                ws.cell(row=row, column=3, value=f"{aed_value:.5f}")
            elif currency == 'USDT':
                usdt_balance = balance
                usdt_value = results['final_usdt_value']
                usdt_in_aed = usdt_value
                ws.cell(row=row, column=3, value=f"{usdt_value:.5f}")
            else:
                ws.cell(row=row, column=3, value="-")

            row += 1

        # Add USDT rate
        ws.cell(row=row, column=1, value="USDT Rate (AED)")
        ws.cell(row=row, column=2, value=f"{results['final_usdt_rate']:.5f}")
        ws.cell(row=row, column=3, value="-")
        row += 1

        # Add subtotal for AED + USDT converted to AED
        subtotal_aed = aed_balance + usdt_in_aed
        ws.cell(row=row, column=1, value="Subtotal (AED + USDT in AED)")
        ws.cell(row=row, column=2, value="-")
        ws.cell(row=row, column=3, value=f"{subtotal_aed:.5f}")
        ws.cell(row=row, column=3).font = Font(bold=True)
        row += 1

        # Add Voo Payment receivables if available
        voo_payment_receivables = results.get('evoucher_aed_received', 0)
        if voo_payment_receivables == 0 and manual_aed_received > 0:
            voo_payment_receivables = manual_aed_received
        ws.cell(row=row, column=1, value="Voo Payment Receivables (AED)")
        ws.cell(row=row, column=2, value=f"{voo_payment_receivables:.5f}")
        ws.cell(row=row, column=3, value=f"{voo_payment_receivables:.5f}")
        row += 1

        # Add total value
        total_final_aed = subtotal_aed + voo_payment_receivables
        ws.cell(row=row, column=1, value="Total Value (AED)")
        ws.cell(row=row, column=2, value="-")
        ws.cell(row=row, column=3, value=f"{total_final_aed:.5f}")
        ws.cell(row=row, column=3).font = Font(bold=True)

        final_balances_end = row

        # Add border to the final balances table
        for r in range(final_balances_start - 1, final_balances_end + 1):
            for c in range(1, 4):  # Extended to include the AED value column
                ws.cell(row=r, column=c).border = thin_border

        # Add Profit Summary section
        row += 2  # Reduce space
        ws[f'A{row}'] = 'PROFIT SUMMARY'
        ws[f'A{row}'].font = Font(size=10, bold=True)
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws[f'A{row}'].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        # Add subtitle explaining the table
        row += 1
        ws[f'A{row}'] = 'Summary of profits from P2P and E-Voucher transactions'
        ws[f'A{row}'].font = Font(size=8, italic=True)
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        row += 1  # Reduce space

        # Add headers for profit summary
        ws.cell(row=row, column=1, value="Profit Type").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Amount (AED)").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        ws.cell(row=row, column=2).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        row += 1
        profit_start = row

        # Add profit data
        ws.cell(row=row, column=1, value="Total P2P Profit")
        ws.cell(row=row, column=2, value=f"{results['total_p2p_profit']:.5f}")
        row += 1

        # Add E-Voucher details
        ws.cell(row=row, column=1, value="Total E-Voucher Profit")
        ws.cell(row=row, column=2, value=f"{results['total_evoucher_profit']:.5f}")
        row += 1

        # Add total profit
        ws.cell(row=row, column=1, value="Total Profit")
        ws.cell(row=row, column=2, value=f"{(results['total_p2p_profit'] + results['total_evoucher_profit']):.5f}")
        ws.cell(row=row, column=2).font = Font(bold=True)

        profit_end = row

        # Add border to the profit summary table
        for r in range(profit_start - 1, profit_end + 1):
            for c in range(1, 3):
                ws.cell(row=r, column=c).border = thin_border

        # Add USDT Sell Transactions Details table
        row += 2  # Reduce space
        ws[f'A{row}'] = 'USDT SELL TRANSACTIONS DETAILS'
        ws[f'A{row}'].font = Font(size=10, bold=True)
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws[f'A{row}'].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        # Add subtitle explaining the table
        row += 1
        ws[f'A{row}'] = 'This table shows detailed calculations for each USDT sell transaction, including the original purchase rates'
        ws[f'A{row}'].font = Font(size=8, italic=True)
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        row += 1  # Reduce space

        # Add Debug Transactions data if available
        if results['debug_transactions'] is not None and not results['debug_transactions'].empty:
            # Create a custom table for USDT sell transactions
            debug_df = results['debug_transactions'].copy()

            # Define custom columns for better readability
            custom_columns = [
                "Date",
                "USDT Amount",
                "Sell Rate (AED)",
                "Total Value (AED)",
                "Original Purchase Rate",
                "Profit (AED)",
                "USDT Used Details"
            ]

            # Add headers
            for c_idx, col_name in enumerate(custom_columns, 1):
                cell = ws.cell(row=row, column=c_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            row += 1
            debug_data_start = row

            # Add data rows
            for _, data_row in debug_df.iterrows():
                # Date
                cell = ws.cell(row=row, column=1, value=data_row.get('date', ''))
                cell.alignment = Alignment(horizontal='left')

                # USDT Amount
                cell = ws.cell(row=row, column=2, value=f"{float(data_row.get('usdt_amount', 0)):.5f}")
                cell.alignment = Alignment(horizontal='right')

                # Sell Rate
                cell = ws.cell(row=row, column=3, value=f"{float(data_row.get('sell_price', 0)):.5f}")
                cell.alignment = Alignment(horizontal='right')

                # Total Value
                cell = ws.cell(row=row, column=4, value=f"{float(data_row.get('total_sell_value', 0)):.5f}")
                cell.alignment = Alignment(horizontal='right')

                # Original Purchase Rate
                cell = ws.cell(row=row, column=5, value=f"{float(data_row.get('original_purchase_rate', 0)):.5f}")
                cell.alignment = Alignment(horizontal='right')

                # Profit
                cell = ws.cell(row=row, column=6, value=f"{float(data_row.get('profit', 0)):.5f}")
                cell.alignment = Alignment(horizontal='right')

                # USDT Used Details - Format as multi-line text with better formatting
                usdt_used = data_row.get('usdt_used', [])
                if isinstance(usdt_used, list) and usdt_used:
                    usdt_details = []
                    for i, (amount, price) in enumerate(usdt_used, 1):
                        # Format with line number for better readability
                        usdt_details.append(f"{i}. {amount:.5f} @ {price:.5f}")

                    # Join with line breaks for multi-line cell
                    usdt_details_text = "\n".join(usdt_details)

                    cell = ws.cell(row=row, column=7, value=usdt_details_text)
                    # Set alignment and wrap text for multi-line display
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    # Set row height to accommodate multiple lines but keep it reasonable
                    line_count = min(len(usdt_details), 5)  # Limit to 5 visible lines max
                    ws.row_dimensions[row].height = 12 * (line_count + 1)
                else:
                    cell = ws.cell(row=row, column=7, value="No details")
                    cell.alignment = Alignment(horizontal='left')

                row += 1

            debug_end_row = row - 1

            # Add border to the table
            for r in range(debug_data_start - 1, debug_end_row + 1):
                for c in range(1, len(custom_columns) + 1):
                    ws.cell(row=r, column=c).border = thin_border



        # Set page setup for A4 printing
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins

        # Set page to A4 size
        ws.page_setup.paperSize = 9  # A4 paper
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToHeight = 0  # Auto
        ws.page_setup.fitToWidth = 1  # Fit to 1 page wide

        # Set minimal margins to maximize content area
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25)

        # Set print title rows (headers)
        ws.print_title_rows = f'1:5'  # Repeat first 5 rows on each page

        # Set column widths optimized for A4 - with fix for merged cells
        # Narrower columns to fit more on a page
        for i in range(1, 20):  # Assuming we won't have more than 20 columns
            column_letter = get_column_letter(i)
            ws.column_dimensions[column_letter].width = 10  # Make columns narrower

        # Set specific widths for certain columns
        ws.column_dimensions['A'].width = 13  # First column (usually dates)
        ws.column_dimensions['B'].width = 20  # Transaction column - wider for wrapped text

        # Set font size smaller for better fit
        from openpyxl.styles import Font
        default_font = Font(name='Arial', size=9)  # Smaller font size

        # Apply smaller font to all cells - with fix for merged cells
        for row in ws.iter_rows():
            for cell in row:
                # Skip merged cells
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue

                if cell.font.size is None or cell.font.size > 9:
                    new_font = Font(
                        name='Arial',
                        size=9,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        color=cell.font.color
                    )
                    cell.font = new_font

        # Save the workbook to the BytesIO object
        wb.save(output)
        output.seek(0)

        return output

    except Exception as e:
        st.error(f"Error exporting to Excel: {str(e)}")
        return None

# Function to combine multiple Excel files
def combine_excel_files(files, preview_only=False, max_preview_rows=100):
    try:
        all_dfs = []
        file_names = []

        # Process each file
        for file in files:
            try:
                # Read Excel file
                df = pd.read_excel(file)

                # Add file name as a column for reference
                file_name = os.path.basename(file.name) if hasattr(file, 'name') else "Unknown"
                df['Source File'] = file_name
                file_names.append(file_name)

                # Add to list of dataframes
                all_dfs.append(df)
            except Exception as e:
                st.error(f"Error reading file {file.name if hasattr(file, 'name') else 'Unknown'}: {str(e)}")
                continue

        if not all_dfs:
            st.error("No valid Excel files were found.")
            return None, None

        # Combine all dataframes
        combined_df = pd.concat(all_dfs, ignore_index=True)

        # If preview only, return a limited number of rows
        if preview_only:
            preview_df = combined_df.head(max_preview_rows)
            return preview_df, file_names

        return combined_df, file_names

    except Exception as e:
        st.error(f"Error combining Excel files: {str(e)}")
        return None, None

# Function to export combined data to Excel
def export_combined_excel(combined_df, file_names):
    try:
        # Create a BytesIO object to store the Excel file
        output = io.BytesIO()

        # Create a workbook
        wb = Workbook()

        # Use the default sheet
        ws = wb.active
        ws.title = "Combined Data"

        # Add title
        ws['A1'] = 'Combined Excel Data'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center')

        # Add subtitle with file names
        row = 2
        ws[f'A{row}'] = f'Files combined: {", ".join(file_names)}'
        ws[f'A{row}'].font = Font(size=10, italic=True)
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        row += 2  # Add some space

        # Add headers
        for c_idx, col_name in enumerate(combined_df.columns, 1):
            cell = ws.cell(row=row, column=c_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        row += 1

        # Add data rows
        for _, data_row in combined_df.iterrows():
            for c_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=c_idx, value=value)

                # Format dates and numbers
                if isinstance(value, (int, float)):
                    cell.number_format = '0.00000'
                    cell.alignment = Alignment(horizontal='right')
                elif isinstance(value, datetime):
                    cell.number_format = 'yyyy-mm-dd'
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='left')

            row += 1

        # Add border to the table
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

        for r in range(4, row):
            for c in range(1, len(combined_df.columns) + 1):
                ws.cell(row=r, column=c).border = thin_border

        # Auto-adjust column widths - with fix for merged cells
        for i, column in enumerate(ws.columns):
            max_length = 0
            column_letter = get_column_letter(i+1)  # Get column letter directly from index

            for cell in column:
                # Skip merged cells
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue

                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length

            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = min(adjusted_width, 30)  # Cap width at 30

        # Save the workbook to the BytesIO object
        wb.save(output)
        output.seek(0)

        return output

    except Exception as e:
        st.error(f"Error exporting to Excel: {str(e)}")
        return None

# Function to analyze E-Voucher transactions
def analyze_evoucher(df, balance_history):
    evoucher_df = df[df['TradeType'] == 'E-Voucher'].copy()

    # Get E-Voucher transactions from balance history for profit calculation
    evoucher_history = balance_history[balance_history['TradeType'] == 'E-Voucher'].copy()

    # Calculate metrics
    total_aed_received = evoucher_df[evoucher_df['Type'] == 'Buy']['Real Amount'].sum()
    total_usdt_sold = evoucher_df[evoucher_df['Type'] == 'Sell']['USDT'].sum()
    avg_aed_egp_rate = evoucher_df['AED/EGP'].mean()

    # Calculate E-Voucher profit
    # For each USDT sold in E-Voucher, calculate its cost in AED
    evoucher_profit = 0

    if not evoucher_history.empty:
        # Get all Buy transactions (AED received)
        buy_transactions = evoucher_history[evoucher_history['Type'] == 'Buy']
        total_aed_received = buy_transactions['Real Amount'].sum()

        # Get all Sell transactions (USDT sold)
        sell_transactions = evoucher_history[evoucher_history['Type'] == 'Sell']

        # Calculate the cost of USDT sold
        total_usdt_cost = 0
        for _, row in sell_transactions.iterrows():
            usdt_amount = row['USDT']
            usdt_rate = row['Current USDT Rate']
            total_usdt_cost += usdt_amount * usdt_rate

        # Calculate profit
        evoucher_profit = total_aed_received - total_usdt_cost

    # Create summary
    summary = {
        'Total AED Received': total_aed_received,
        'Total USDT Sold': total_usdt_sold,
        'Average AED/EGP Rate': avg_aed_egp_rate,
        'E-Voucher Profit': evoucher_profit
    }

    # Create rate chart
    rate_fig = px.line(
        evoucher_df,
        x='Date',
        y='AED/EGP',
        title='AED/EGP Exchange Rate Over Time'
    )

    return summary, rate_fig, evoucher_profit

# Function to save settings to a JSON file
def save_settings(settings):
    try:
        with open('p2p_settings.json', 'w') as f:
            json.dump(settings, f)
        return True
    except Exception as e:
        st.error(f"Error saving settings: {str(e)}")
        return False

# Function to load settings from a JSON file
def load_settings():
    try:
        if os.path.exists('p2p_settings.json'):
            with open('p2p_settings.json', 'r') as f:
                return json.load(f)
    except Exception as e:
        st.error(f"Error loading settings: {str(e)}")
    return {}

# Function to transfer final balances to initial settings
def transfer_final_balances_to_settings(results):
    try:
        # Get current settings
        settings = load_settings()

        # Update initial balances with final balances
        settings['initial_balances'] = results['final_balances']

        # Update initial USDT rate with final USDT rate
        settings['initial_usdt_rate'] = results['final_usdt_rate']

        # Save updated settings
        if save_settings(settings):
            return True
        else:
            return False
    except Exception as e:
        st.error(f"Error transferring balances: {str(e)}")
        return False

# Main application
def main():
    # Initialize session state for settings if not already done
    if 'settings_loaded' not in st.session_state:
        st.session_state.settings_loaded = True
        # Load saved settings
        saved_settings = load_settings()

        # Initialize session state with saved settings or defaults
        if 'initial_balances' not in st.session_state:
            st.session_state.initial_balances = saved_settings.get('initial_balances', {})
        if 'initial_usdt_rate' not in st.session_state:
            st.session_state.initial_usdt_rate = saved_settings.get('initial_usdt_rate', 3.67)
        if 'evoucher_aed_received' not in st.session_state:
            st.session_state.evoucher_aed_received = saved_settings.get('evoucher_aed_received', 0.0)
        if 'additional_currencies' not in st.session_state:
            st.session_state.additional_currencies = saved_settings.get('additional_currencies', [])

    # Sidebar for initial balances
    st.sidebar.subheader("Enter Initial Balances")

    # Default currencies
    default_currencies = ['AED', 'EGP', 'USDT']

    # Ensure default currencies exist in session state
    for currency in default_currencies:
        if currency not in st.session_state.initial_balances:
            st.session_state.initial_balances[currency] = 0.0

    # Initial balances
    initial_balances = {}
    for currency in default_currencies:
        initial_balances[currency] = st.sidebar.number_input(
            f"Initial {currency} Balance",
            value=st.session_state.initial_balances.get(currency, 0.0),
            step=0.01,
            format="%.15f",
            key=f"balance_{currency}"
        )
        # Update session state when value changes
        st.session_state.initial_balances[currency] = initial_balances[currency]

    # Initial USDT rate
    st.sidebar.subheader("Initial USDT Rate")
    initial_usdt_rate = st.sidebar.number_input(
        "Initial USDT Rate (AED)",
        value=st.session_state.initial_usdt_rate,
        step=0.01,
        format="%.15f",
        help="This is the rate at which your initial USDT balance was acquired",
        key="usdt_rate"
    )
    # Update session state
    st.session_state.initial_usdt_rate = initial_usdt_rate

    # E-Voucher AED received
    st.sidebar.subheader("E-Voucher Settings")
    evoucher_aed_received = st.sidebar.number_input(
        "Total AED Received by Voo Payment",
        value=st.session_state.evoucher_aed_received,
        step=100.0,
        format="%.15f",
        help="Enter the total AED amount received by Voo Payment from workers for E-Voucher transactions (receivable from Voo Payment)",
        key="evoucher_aed"
    )
    # Update session state
    st.session_state.evoucher_aed_received = evoucher_aed_received

    # Option to add more currencies
    st.sidebar.subheader("Add Additional Currency")
    new_currency = st.sidebar.text_input("Currency Code (e.g., USD)", key="new_currency")

    # Add additional currencies from session state
    for currency in st.session_state.additional_currencies:
        if currency not in initial_balances and currency not in default_currencies:
            initial_balances[currency] = st.sidebar.number_input(
                f"Initial {currency} Balance",
                value=st.session_state.initial_balances.get(currency, 0.0),
                step=0.01,
                format="%.15f",
                key=f"balance_{currency}"
            )
            # Update session state
            st.session_state.initial_balances[currency] = initial_balances[currency]

    # Add new currency if entered
    if new_currency and new_currency not in initial_balances:
        initial_balances[new_currency] = st.sidebar.number_input(
            f"Initial {new_currency} Balance",
            value=0.0,
            step=0.01,
            format="%.15f",
            key=f"balance_{new_currency}"
        )
        # Update session state
        st.session_state.initial_balances[new_currency] = initial_balances[new_currency]
        if new_currency not in st.session_state.additional_currencies:
            st.session_state.additional_currencies.append(new_currency)

    # Save button for settings
    if st.sidebar.button("💾 Save Settings", help="Save current settings for future sessions"):
        settings = {
            'initial_balances': st.session_state.initial_balances,
            'initial_usdt_rate': st.session_state.initial_usdt_rate,
            'evoucher_aed_received': st.session_state.evoucher_aed_received,
            'additional_currencies': st.session_state.additional_currencies
        }
        if save_settings(settings):
            st.sidebar.success("Settings saved successfully!")
        else:
            st.sidebar.error("Failed to save settings.")

    # Create main tabs
    main_tabs = st.tabs(["تتبع المعاملات", "دمج الملفات"])

    # Tab for transaction tracking
    with main_tabs[0]:
        # File uploader
        st.subheader("Upload Transaction File")

        # Initialize last_file_path in session state if not present
        if 'last_file_path' not in st.session_state:
            st.session_state.last_file_path = saved_settings.get('last_file_path', '')

        # Show last used file path if available
        if st.session_state.last_file_path:
            st.info(f"Last used file: {os.path.basename(st.session_state.last_file_path)}")

            # Option to use the last file
            col1, col2 = st.columns([1, 3])
            with col1:
                if os.path.exists(st.session_state.last_file_path) and st.button("📂 Use Last File"):
                    uploaded_file = open(st.session_state.last_file_path, "rb")
                else:
                    uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "csv"])
        else:
            uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "csv"])

        # Save file path if a file is uploaded
        if uploaded_file is not None and hasattr(uploaded_file, 'name'):
            try:
                # Try to get the file path
                file_path = uploaded_file.name
                if os.path.exists(file_path):
                    st.session_state.last_file_path = file_path
                    # Update settings
                    settings = {
                        'initial_balances': st.session_state.initial_balances,
                        'initial_usdt_rate': st.session_state.initial_usdt_rate,
                        'evoucher_aed_received': st.session_state.evoucher_aed_received,
                        'additional_currencies': st.session_state.additional_currencies,
                        'last_file_path': file_path
                    }
                    save_settings(settings)
            except:
                pass  # Ignore if we can't get the file path

        if uploaded_file is not None:
            # Process data
            results = process_data(uploaded_file, initial_balances, initial_usdt_rate)

            # Update E-Voucher profit with manually entered AED amount
            if results and evoucher_aed_received > 0:
                # Get all E-Voucher sell transactions
                evoucher_df = results['transactions'][
                    (results['transactions']['TradeType'] == 'E-Voucher') &
                    (results['transactions']['Type'] == 'Sell')
                ]

                if not evoucher_df.empty:
                    # Calculate the cost of USDT sold
                    total_usdt_cost = 0
                    for _, row in evoucher_df.iterrows():
                        usdt_amount = row['USDT']
                        # Find the corresponding row in balance history
                        idx = results['balance_history'][results['balance_history']['Date'] == row['Date']].index[0]
                        usdt_rate = results['balance_history'].loc[idx, 'Current USDT Rate']
                        total_usdt_cost += usdt_amount * usdt_rate

                    # Calculate profit
                    results['total_evoucher_profit'] = evoucher_aed_received - total_usdt_cost

            if results:
                # Create tabs
                tabs = st.tabs(["Dashboard", "Cash Flow", "Transactions", "P2P Analysis", "E-Voucher Analysis", "Debug"])

            # Dashboard tab
            with tabs[0]:
                # Add export to Excel button and transfer balances button at the top of the dashboard
                col1, col2, col3 = st.columns([2, 1, 1])
                with col2:
                    # Create Excel export button
                    excel_data = export_to_excel(results, evoucher_aed_received)
                    # Make sure evoucher_aed_received is included in results
                    results['evoucher_aed_received'] = evoucher_aed_received
                    if excel_data:
                        # Get transaction date range for the filename
                        file_date = "no_data"
                        if not results['cash_flow'].empty:
                            try:
                                # Get only the first date of transactions
                                first_date = results['cash_flow']['Date'].iloc[0] if 'Date' in results['cash_flow'].columns else ""

                                # Make sure date is string
                                first_date = str(first_date)

                                # Extract just the date part (remove time if present)
                                if ' ' in first_date:
                                    first_date = first_date.split(' ')[0]

                                # Format date for filename (YYYY-MM-DD format)
                                # Convert from DD/MM/YYYY to YYYY-MM-DD if needed
                                if '/' in first_date:
                                    parts = first_date.split('/')
                                    if len(parts) == 3:
                                        # Assuming DD/MM/YYYY format
                                        day, month, year = parts
                                        file_date = f"{year}-{month}-{day}"
                                    else:
                                        file_date = first_date.replace('/', '-')
                                else:
                                    file_date = first_date

                            except Exception as e:
                                # If there's any error, use a default name
                                from datetime import datetime
                                today = datetime.now().strftime("%Y-%m-%d")
                                file_date = today
                                st.warning(f"Could not format date for filename: {str(e)}")

                        # Get total profit for filename
                        total_profit = results['total_p2p_profit'] + results['total_evoucher_profit']
                        profit_str = f"{total_profit:.2f}AED"

                        # Create a professional filename with date and profit
                        excel_filename = f"P2P_Report_{file_date}_Profit_{profit_str}.xlsx"

                        st.download_button(
                            label="📊 Export All Data to Excel",
                            data=excel_data,
                            file_name=excel_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            help="Export all data from all tabs to a single Excel file with multiple sheets"
                        )

                with col3:
                    # Create transfer balances button
                    if st.button(
                        "🔄 Transfer Closing Balances",
                        help="Transfer closing balances to become opening balances for the next session"
                    ):
                        if transfer_final_balances_to_settings(results):
                            st.success("Closing balances successfully transferred to opening balances!")
                        else:
                            st.error("Failed to transfer balances.")

                # Summary metrics - using custom styled cards
                st.markdown(f"""
                <div class="profit-card">
                    <div class="profit-title">Total P2P Profit (AED)</div>
                    <div class="profit-value">{results['total_p2p_profit']:.5f}</div>
                </div>
                """, unsafe_allow_html=True)

                st.markdown(f"""
                <div class="profit-card">
                    <div class="profit-title">Total E-Voucher Profit (AED)</div>
                    <div class="profit-value">{results['total_evoucher_profit']:.5f}</div>
                    <div class="profit-subtitle">Receivable from Voo Payment</div>
                </div>
                """, unsafe_allow_html=True)

                st.markdown(f"""
                <div class="profit-card">
                    <div class="profit-title">Total Profit (AED)</div>
                    <div class="profit-value">{results['total_p2p_profit'] + results['total_evoucher_profit']:.5f}</div>
                </div>
                """, unsafe_allow_html=True)

                # Explanation of profit calculation
                st.info("""
                **Profit Calculation:**
                - **P2P Profit**: Calculated from P2P transactions only, using the FIFO method (first in, first out)
                - **E-Voucher Profit**: Calculated as (AED received by Voo Payment - cost of USDT sold)
                  * Note: This profit is receivable from Voo Payment, as they receive the AED from workers
                - **Total Profit**: Sum of P2P and E-Voucher profits
                """)


                # Final balances
                st.subheader("Final Balances")

                # Calculate total value in AED (AED + USDT + Voo Payment receivables)
                owned_assets_value = (
                    results['final_balances'].get('AED', 0) +
                    results['final_usdt_value']
                )

                # Get Voo Payment receivables
                voo_payment_receivables = evoucher_aed_received if evoucher_aed_received > 0 else 0

                # Calculate total value including receivables
                total_value_aed = owned_assets_value + voo_payment_receivables

                # Display each currency in a separate card
                for currency, balance in results['final_balances'].items():
                    if currency == 'USDT':
                        st.markdown(f"""
                        <div class="currency-card">
                            <div class="currency-title">{currency}</div>
                            <div class="currency-value">{balance:.5f}</div>
                            <div class="currency-subtitle">≈ {results['final_usdt_value']:.5f} AED</div>
                        </div>
                        """, unsafe_allow_html=True)
                    elif currency == 'EGP':
                        st.markdown(f"""
                        <div class="currency-card">
                            <div class="currency-title">{currency}</div>
                            <div class="currency-value">{balance:.5f}</div>
                            <div class="currency-subtitle">Not included in total value</div>
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        st.markdown(f"""
                        <div class="currency-card">
                            <div class="currency-title">{currency}</div>
                            <div class="currency-value">{balance:.5f}</div>
                        </div>
                        """, unsafe_allow_html=True)

                # Display USDT rate
                st.markdown(f"""
                <div class="currency-card">
                    <div class="currency-title">USDT Rate</div>
                    <div class="currency-value">{results['final_usdt_rate']:.5f} AED</div>
                </div>
                """, unsafe_allow_html=True)

                # Display owned assets value
                st.markdown(f"""
                <div class="total-value-card" style="background-color: #f5f5f5; margin-bottom: 10px;">
                    <div class="currency-title">Owned Assets Value (AED)</div>
                    <div class="currency-value">{owned_assets_value:.5f}</div>
                    <div class="currency-subtitle">Sum of AED and USDT value that you directly own</div>
                </div>
                """, unsafe_allow_html=True)

                # Display Voo Payment receivables
                st.markdown(f"""
                <div class="total-value-card" style="background-color: #f0f8ff; margin-bottom: 10px;">
                    <div class="currency-title">Voo Payment Receivables (AED)</div>
                    <div class="currency-value">{voo_payment_receivables:.5f}</div>
                    <div class="currency-subtitle">Amount receivable from Voo Payment</div>
                </div>
                """, unsafe_allow_html=True)

                # Display total value
                st.markdown(f"""
                <div class="total-value-card" style="background-color: #e6ffe6; border: 2px solid #4CAF50;">
                    <div class="currency-title">Total Value (AED)</div>
                    <div class="currency-value">{total_value_aed:.5f}</div>
                    <div class="currency-subtitle">Sum of Owned Assets ({owned_assets_value:.5f}) + Voo Payment Receivables ({voo_payment_receivables:.5f})</div>
                </div>
                """, unsafe_allow_html=True)

                # Charts
                st.subheader("Balance History")
                balance_chart = create_balance_chart(results['balance_history'], results['final_balances'].keys())
                st.plotly_chart(balance_chart, use_container_width=True)

                st.subheader("Profit History")
                profit_chart = create_profit_chart(results['profit_history'])
                st.plotly_chart(profit_chart, use_container_width=True)

            # Cash Flow tab
            with tabs[1]:
                st.subheader("Cash Flow Analysis")

                # Display cash flow table
                st.dataframe(results['cash_flow'], use_container_width=True)

                # Create cash flow chart
                st.subheader("Total Value Over Time")

                fig = go.Figure()

                fig.add_trace(
                    go.Scatter(
                        x=results['cash_flow']['Date'],
                        y=results['cash_flow']['Total Value (AED)'],
                        mode='lines+markers',
                        name='Total Value (AED)',
                        line=dict(width=3, color='green')
                    )
                )

                fig.update_layout(
                    xaxis_title='Date',
                    yaxis_title='Value (AED)',
                    height=400
                )

                st.plotly_chart(fig, use_container_width=True)

                # Download button
                csv = results['cash_flow'].to_csv(index=False)
                st.download_button(
                    label="Download Cash Flow as CSV",
                    data=csv,
                    file_name="cash_flow.csv",
                    mime="text/csv"
                )

            # Transactions tab
            with tabs[2]:
                st.subheader("All Transactions")
                st.dataframe(results['balance_history'], use_container_width=True)

                # Download button
                csv = results['balance_history'].to_csv(index=False)
                st.download_button(
                    label="Download Transactions as CSV",
                    data=csv,
                    file_name="transactions_with_balances.csv",
                    mime="text/csv"
                )

            # P2P Analysis tab
            with tabs[3]:
                st.subheader("P2P Transactions Analysis")

                # Filter P2P transactions
                p2p_df = results['balance_history'][results['balance_history']['TradeType'] == 'P2P'].copy()

                # P2P metrics
                p2p_buy = p2p_df[p2p_df['Type'] == 'Buy']
                p2p_sell = p2p_df[p2p_df['Type'] == 'Sell']

                col1, col2, col3 = st.columns(3)

                with col1:
                    st.metric("Total USDT Bought", f"{p2p_buy['USDT'].sum():.15f}")

                with col2:
                    st.metric("Total USDT Sold", f"{p2p_sell['USDT'].sum():.15f}")

                with col3:
                    st.metric("P2P Profit (AED)", f"{results['total_p2p_profit']:.15f}")

                # P2P transactions by currency
                st.subheader("P2P Transactions by Currency")

                currency_counts = p2p_df['Currency'].value_counts().reset_index()
                currency_counts.columns = ['Currency', 'Count']

                fig = px.pie(
                    currency_counts,
                    values='Count',
                    names='Currency',
                    title='Distribution of P2P Transactions by Currency'
                )
                st.plotly_chart(fig, use_container_width=True)

                # P2P transactions table
                st.subheader("P2P Transactions")
                st.dataframe(p2p_df, use_container_width=True)

            # E-Voucher Analysis tab
            with tabs[4]:
                st.subheader("E-Voucher Transactions Analysis")

                # Filter E-Voucher transactions
                evoucher_df = results['balance_history'][results['balance_history']['TradeType'] == 'E-Voucher'].copy()

                if not evoucher_df.empty:
                    # Get E-Voucher analysis
                    evoucher_summary, rate_chart, _ = analyze_evoucher(results['transactions'], results['balance_history'])

                    # Display metrics
                    col1, col2, col3 = st.columns(3)

                    # If manual AED amount was entered, use that instead
                    aed_received = evoucher_aed_received if evoucher_aed_received > 0 else evoucher_summary['Total AED Received']

                    with col1:
                        st.metric("Total AED Received", f"{aed_received:.15f}",
                                 f"{'(Manual Entry)' if evoucher_aed_received > 0 else ''}")

                    with col2:
                        st.metric("Total USDT Sold", f"{evoucher_summary['Total USDT Sold']:.15f}")

                    with col3:
                        st.metric("E-Voucher Profit (AED)", f"{results['total_evoucher_profit']:.15f}")

                    # Explanation of E-Voucher profit calculation
                    st.info(f"""
                    **E-Voucher Profit Calculation:**
                    1. Total AED received by Voo Payment from workers: {aed_received:.15f} AED {' (Manual Entry)' if evoucher_aed_received > 0 else ''}
                    2. Minus the cost of USDT sold: {evoucher_summary['Total USDT Sold']:.15f} USDT at average rate
                    3. Total cost of USDT: {(results['total_evoucher_profit'] - aed_received) * -1:.15f} AED
                    4. Profit: {results['total_evoucher_profit']:.15f} AED

                    **Important Notes:**
                    - The AED amount is received by Voo Payment, not by you directly. This amount is receivable from Voo Payment.
                    - The EGP sent to families is not counted as an asset since it's transferred out.
                    """)

                    # Display exchange rate chart
                    st.subheader("AED/EGP Exchange Rate")
                    st.plotly_chart(rate_chart, use_container_width=True)

                    # E-Voucher transactions table
                    st.subheader("E-Voucher Transactions")
                    st.dataframe(evoucher_df, use_container_width=True)
                else:
                    st.info("No E-Voucher transactions found in the uploaded file.")

            # Debug tab
            with tabs[5]:
                st.subheader("Debug Information")

                if results['debug_transactions'] is not None:
                    st.write("This tab shows detailed calculations for each sell transaction.")

                    # Display debug transactions
                    for i, row in results['debug_transactions'].iterrows():
                        with st.expander(f"Transaction {i+1}: Sell {row['usdt_amount']:.2f} USDT on {row['date']}"):
                            st.write(f"**Sell Details:**")
                            st.write(f"- Amount: {row['usdt_amount']:.15f} USDT")
                            st.write(f"- Sell Price: {row['sell_price']:.15f} AED")
                            st.write(f"- Total Sell Value: {row['total_sell_value']:.15f} AED")

                            st.write("**USDT Used (FIFO method):**")
                            for j, (amount, price) in enumerate(row['usdt_used']):
                                st.write(f"- Batch {j+1}: {amount:.15f} USDT @ {price:.15f} AED = {amount * price:.15f} AED")

                            st.write(f"**Total Cost Value:** {row['total_cost_value']:.15f} AED")
                            st.write(f"**Original Purchase Rate:** {row['original_purchase_rate']:.15f} AED")
                            st.write(f"**Profit:** {row['profit']:.15f} AED")

                            st.write("**Calculation Details:**")
                            for line in row['debug_info']:
                                st.write(f"- {line}")
                else:
                    st.info("No debug information available.")

    # Tab for Excel Files Merger
    with main_tabs[1]:
        st.subheader("دمج ملفات Excel متعددة")

        st.markdown("""
        <div style="background-color: #f0f8ff; padding: 15px; border-radius: 10px; margin-bottom: 20px;">
            <h4 style="color: #1E88E5;">تعليمات الاستخدام</h4>
            <p>هذه الأداة تسمح لك بدمج عدة ملفات Excel في ملف واحد مع الحفاظ على عناوين الأعمدة.</p>
            <ol>
                <li>قم بتحميل ملفات Excel المتعددة التي تريد دمجها</li>
                <li>سيتم عرض معاينة للبيانات المدمجة</li>
                <li>اضغط على زر "تصدير البيانات المدمجة إلى Excel" لتحميل الملف النهائي</li>
            </ol>
        </div>
        """, unsafe_allow_html=True)

        # File uploader for multiple files
        uploaded_files = st.file_uploader("اختر ملفات Excel للدمج", type=["xlsx", "xls"], accept_multiple_files=True)

        if uploaded_files:
            # Show number of files uploaded
            st.success(f"تم تحميل {len(uploaded_files)} ملفات")

            # List file names
            with st.expander("عرض أسماء الملفات المحملة"):
                for i, file in enumerate(uploaded_files):
                    st.write(f"{i+1}. {file.name}")

            # Combine files and show preview
            preview_df, file_names = combine_excel_files(uploaded_files, preview_only=True, max_preview_rows=100)

            if preview_df is not None:
                st.subheader("معاينة البيانات المدمجة")
                st.dataframe(preview_df, use_container_width=True)

                # Show column information
                with st.expander("معلومات الأعمدة"):
                    st.write(f"عدد الأعمدة: {len(preview_df.columns)}")
                    st.write("أسماء الأعمدة:")
                    for col in preview_df.columns:
                        st.write(f"- {col}")

                # Export button
                st.subheader("تصدير البيانات المدمجة")

                # Get full combined data (not just preview)
                full_df, _ = combine_excel_files(uploaded_files, preview_only=False)

                if full_df is not None:
                    # Create Excel export
                    excel_data = export_combined_excel(full_df, file_names)

                    if excel_data:
                        # Create filename with current date
                        from datetime import datetime
                        today = datetime.now().strftime("%Y-%m-%d")
                        excel_filename = f"Combined_Excel_Data_{today}.xlsx"

                        # Download button
                        st.download_button(
                            label="📊 تصدير البيانات المدمجة إلى Excel",
                            data=excel_data,
                            file_name=excel_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            help="تحميل جميع البيانات المدمجة في ملف Excel واحد"
                        )

                        # Show statistics
                        st.subheader("إحصائيات البيانات المدمجة")
                        st.write(f"إجمالي عدد الصفوف: {len(full_df)}")
                        st.write(f"إجمالي عدد الأعمدة: {len(full_df.columns)}")

                        # Show data types
                        with st.expander("أنواع البيانات"):
                            st.write(full_df.dtypes)
        else:
            st.info("الرجاء تحميل ملفات Excel للدمج")

if __name__ == "__main__":
    main()
